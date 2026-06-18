"""
EduTrack — Student Management Routes
"""
import io
import logging
import re

from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, current_app, jsonify)
from flask_login import login_required, current_user
from openpyxl import load_workbook

from app.extensions import db
from app.models import RegisteredStudent, AttendanceRecord

logger = logging.getLogger(__name__)
students_bp = Blueprint("students", __name__, url_prefix="/students")

REQUIRED_COLUMNS = {"Full Name", "Email", "Student ID"}


def _parse_excel(file_stream) -> tuple[list[dict], list[str]]:
    """
    Parse an Excel file stream.
    Returns (valid_rows, errors)
    """
    try:
        wb = load_workbook(filename=file_stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        return [], [f"Failed to read Excel file: {exc}"]

    headers = []
    rows = []
    errors = []
    email_regex = re.compile(r"^[^@]+@[^@]+\.[^@]+$")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else "" for h in row]
            missing = REQUIRED_COLUMNS - set(headers)
            if missing:
                return [], [f"Missing required columns: {', '.join(missing)}"]
            continue

        if all(cell is None for cell in row):
            continue  # skip empty rows

        row_dict = {headers[j]: str(row[j]).strip() if row[j] is not None else ""
                    for j in range(len(headers))}

        row_errors = []
        for col in REQUIRED_COLUMNS:
            if not row_dict.get(col):
                row_errors.append(f"Row {i + 1}: Missing '{col}'")

        email = row_dict.get("Email", "")
        if email and not email_regex.match(email):
            row_errors.append(f"Row {i + 1}: Invalid email '{email}'")

        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(row_dict)

    return rows, errors


@students_bp.route("/")
@login_required
def list_students():
    active_course_id = session.get("active_course_id")
    query = RegisteredStudent.query.filter_by(
        professor_id=current_user.id, is_active=True, course_id=active_course_id
    )
    students = query.order_by(RegisteredStudent.full_name).all()

    return render_template("students/list.html",
                           students=students)


@students_bp.route("/upload", methods=["GET", "POST"])
@login_required
def upload_students():
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file provided.", "danger")
            return redirect(request.url)

        file = request.files["file"]
        if not file.filename.endswith((".xlsx", ".xls")):
            flash("Please upload a valid Excel file (.xlsx or .xls).", "danger")
            return redirect(request.url)

        file_stream = io.BytesIO(file.read())
        rows, parse_errors = _parse_excel(file_stream)

        added = 0
        duplicates = 0
        errors = list(parse_errors)

        active_course_id = session.get("active_course_id")
        active_course_code = session.get("active_course_code", "ALL")

        for row in rows:
            email = row["Email"].lower()

            existing = RegisteredStudent.query.filter_by(
                email=email,
                course_id=active_course_id,
                professor_id=current_user.id,
            ).first()

            if existing:
                duplicates += 1
                continue

            try:
                student = RegisteredStudent(
                    full_name=row["Full Name"],
                    email=email,
                    student_id=row["Student ID"],
                    course_code=active_course_code,
                    course_id=active_course_id,
                    professor_id=current_user.id,
                )
                db.session.add(student)
                db.session.flush()
                added += 1
            except Exception as exc:
                db.session.rollback()
                errors.append(f"Failed to add {row.get('Full Name', '?')}: {exc}")

        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            flash(f"Database error: {exc}", "danger")
            return redirect(request.url)

        return render_template(
            "students/upload.html",
            upload_done=True,
            added=added,
            duplicates=duplicates,
            errors=errors,
        )

    return render_template("students/upload.html", upload_done=False)


@students_bp.route("/<int:student_id>")
@login_required
def student_detail(student_id):
    student = RegisteredStudent.query.filter_by(
        id=student_id, professor_id=current_user.id
    ).first_or_404()

    records = AttendanceRecord.query.filter_by(
        student_id=student_id
    ).order_by(AttendanceRecord.created_at.desc()).all()

    return render_template("students/detail.html",
                           student=student, records=records)


@students_bp.route("/<int:student_id>/link", methods=["POST"])
@login_required
def link_student(student_id):
    """Manually link a Zoom display name to this student (trains model)."""
    data = request.get_json()
    zoom_name = data.get("zoom_name", "").strip()
    zoom_email = data.get("zoom_email", "").strip()
    record_id = data.get("record_id")
    confirmed = data.get("confirmed", True)  # True = same person

    if not zoom_name:
        return jsonify({"error": "zoom_name is required"}), 400

    student = RegisteredStudent.query.filter_by(
        id=student_id, professor_id=current_user.id
    ).first_or_404()

    # Update the attendance record if provided
    if record_id:
        record = AttendanceRecord.query.get(record_id)
        if record and record.session.professor_id == current_user.id:
            record.needs_review = False
            record.match_method = "manual"
            record.match_confidence_score = 1.0 if confirmed else 0.0
            record.is_present = confirmed and (record.total_duration_seconds or 0) >= \
                                 current_app.config.get("SESSION_DURATION_THRESHOLD", 3600)
            db.session.commit()

    # Add training pair for model improvement
    label = 1 if confirmed else 0
    try:
        from app.services.matching import add_training_pair
        add_training_pair(
            zoom_name=zoom_name,
            zoom_email=zoom_email,
            reg_name=student.full_name,
            reg_email=student.email,
            label=label,
            retrain=False,  # Retrain will happen nightly
        )
    except Exception as exc:
        logger.warning(f"Failed to add training pair: {exc}")

    return jsonify({"success": True, "message": "Link saved successfully."})


@students_bp.route("/reset-data", methods=["POST"])
@login_required
def reset_data():
    """Clear all students, sessions, events, and reports for current professor."""
    try:
        from app.models import ZoomSession, RegisteredStudent, AttendanceRecord, ParticipantEvent, AttendanceReport
        
        # Get IDs for current professor's students and sessions
        student_ids = [s.id for s in RegisteredStudent.query.filter_by(professor_id=current_user.id).all()]
        session_ids = [s.id for s in ZoomSession.query.filter_by(professor_id=current_user.id).all()]
        
        deleted_records = 0
        deleted_events = 0
        deleted_reports = 0
        
        # Explicitly delete child records from SQLite to prevent orphaned rows
        if student_ids:
            deleted_records = AttendanceRecord.query.filter(
                AttendanceRecord.student_id.in_(student_ids)
            ).delete(synchronize_session=False)
            
        if session_ids:
            deleted_events = ParticipantEvent.query.filter(
                ParticipantEvent.session_id.in_(session_ids)
            ).delete(synchronize_session=False)
            
            deleted_reports = AttendanceReport.query.filter(
                AttendanceReport.session_id.in_(session_ids)
            ).delete(synchronize_session=False)
        
        deleted_sessions = ZoomSession.query.filter_by(professor_id=current_user.id).delete(synchronize_session=False)
        deleted_students = RegisteredStudent.query.filter_by(professor_id=current_user.id).delete(synchronize_session=False)
        
        db.session.commit()
        flash(f"Successfully cleared all data: {deleted_sessions} sessions, {deleted_students} students, and {deleted_records} attendance records.", "success")
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Reset data failed: {exc}")
        flash("Failed to clear data. Please try again.", "danger")
        
    return redirect(url_for("dashboard.dashboard"))


@students_bp.route("/export-csv")
@login_required
def export_students_csv():
    """Export the student roster with their overall attendance rates to a CSV file."""
    import io
    import csv
    from flask import Response
    
    # Query students
    students = RegisteredStudent.query.filter_by(
        professor_id=current_user.id, is_active=True
    ).order_by(RegisteredStudent.full_name).all()
    
    # Build CSV in memory synchronously to prevent closed session issues during streaming
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    
    # Write header
    writer.writerow(["Student", "Email", "Student ID", "Course", "Attendance"])
    
    # Write rows
    for s in students:
        attendance_str = f"{s.attendance_rate}%" if s.has_records else "No sessions yet"
        writer.writerow([
            s.full_name,
            s.email,
            s.student_id,
            s.course_code,
            attendance_str
        ])
        
    csv_data = output.getvalue()
    output.close()
    
    response = Response(csv_data, mimetype="text/csv")
    response.headers.set("Content-Disposition", "attachment", filename="student_attendance_roster.csv")
    return response

